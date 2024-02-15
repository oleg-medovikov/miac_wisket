from pandas.io.formats.style import Styler
from openpyxl.styles import Alignment, Border, Side
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter


column_widths = {1: 10, 2: 60, 3: 20, 4: 25, 5: 15}


def write_file(df: Styler, filename: str):
    # Сохраняем стилизованный DataFrame в файл Excel
    df.to_excel(filename, sheet_name="tabel", engine="openpyxl")

    # Загружаем рабочий лист
    wb = load_workbook(filename)
    ws = wb["tabel"]

    # перенос текста во второй колонке
    for cell in ws["B"]:
        cell.alignment = Alignment(
            wrap_text=True, vertical="center", horizontal="center"
        )

    # ширина столбцов
    for k, v in column_widths.items():
        ws.column_dimensions[get_column_letter(k)].width = v

    # Устанавливаем фиксированную ширину для остальных колонок
    for i in range(len(column_widths) + 1, ws.max_column + 1):
        ws.column_dimensions[get_column_letter(i)].width = 17  # Фиксированная ширина 20

    # добавляем границы
    # Определяем стиль границ
    border_top = Border(top=Side(style="thin"))
    # border_bottom = Border(bottom=Side(style="thin"))

    # Применяем стиль границ к первым 10 строкам
    for row in ws.iter_rows(
        min_row=5,
        max_row=ws.max_row,
        min_col=2,
        max_col=ws.max_column,
    ):
        for cell in row:
            if cell.row % 4 == 0:  # Если строка четная
                cell.border = border_top
            # else:  # Если строка нечетная
            #    cell.border = border_bottom

    # удаляем столбец
    ws.delete_cols(5, 1)
    # Закрепляем первые пять столбцов и две строчки
    ws.freeze_panes = "E4"
    # Сохраняем изменения
    wb.save(filename)
